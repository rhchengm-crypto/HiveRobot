from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT = Path("HP60C_网页抓取启动流程.xlsx")


def style_sheet(ws, widths, row_height=58):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(name="Arial", color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if not cell.font.bold:
                cell.font = Font(name="Arial", size=10)
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = row_height
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False


def main():
    wb = Workbook()
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")

    ws = wb.active
    ws.title = "总流程"
    ws.append(["步骤", "任务", "位置", "命令 / 操作", "注意事项", "成功标志"])
    steps = [
        [1, "Orin 开机并登录", "Orin", "开机后登录 nvidia 用户，确认 HP60C、机械臂 CAN/电源已连接。", "", "只做硬件准备"],
        [2, "启动 HP60C ROS 驱动", "Orin 终端 1", "cd ~/ascam_ws\nsource ~/.bashrc\nroslaunch ascamera hp60c.launch", "保持终端不关闭", "正常会启动 ROS master 和相机 topic"],
        [3, "首次配置免密 sudo（只需一次）", "Orin 终端 2", "cd ~/hive_robot/DM_Control_Python\nsudo bash install_hive_robot_sudoers.sh /home/nvidia/hive_robot/DM_Control_Python/left_arm_controller.py\nsudo -n python3 /home/nvidia/hive_robot/DM_Control_Python/left_arm_controller.py status", "只允许 left_arm_controller.py 免密，不是全局免密", "第二条 status 不要求密码即成功"],
        [4, "启动网页推流和控制服务", "Orin 终端 2", "cd ~/hive_robot/DM_Control_Python\npython3 hp60c_stream_server.py --scan-windows --window-w 80 --window-h 85 --window-step 20 --min-area 30 --max-area 6000 --aim-y-frac 0.25 --target-depth-percentile 75 --port 8090 --enable-controller-execute", "服务启动后不要关闭该终端", "看到 Controller execute enabled: True"],
        [5, "Windows 打开网页", "Windows 浏览器", "http://<orin-ip>:8090/", "如果不知道 IP，在 Orin 上用 hostname -I 查看", "能看到实时画面、紫框、黄圈"],
        [6, "先验证姿态", "Windows 网页", "勾选 skip claw -> 点击 执行当前目标", "机械臂会运动，但不会闭爪", "确认侧摆、手肘、手腕姿态正确"],
        [7, "正式抓取", "Windows 网页", "取消 skip claw -> 点击 执行当前目标", "机械臂会闭爪抓取", "观察是否抓到目标物"],
        [8, "记录结果", "Windows 网页", "点击 记录完美 / 记录成功 / 记录失败，并输入备注", "失败备注写清楚：手腕/侧摆/肘/前摆怎么偏", "记录写入 data/hp60c_web_runs.jsonl"],
        [9, "回 Home", "Windows 网页", "点击 Home", "执行 sudo -n python3 left_arm_controller.py home", "机械臂安全回 home"],
    ]
    for row in steps:
        ws.append(row)
    style_sheet(ws, [8, 22, 18, 78, 42, 34], 78)
    for row in range(2, ws.max_row + 1):
        fill = warn_fill if ws.cell(row, 1).value in (3, 6, 7) else ok_fill if ws.cell(row, 1).value in (8, 9) else None
        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = fill

    ws = wb.create_sheet("命令速查")
    ws.append(["用途", "命令", "什么时候用"])
    commands = [
        ["启动 HP60C ROS", "cd ~/ascam_ws\nsource ~/.bashrc\nroslaunch ascamera hp60c.launch", "每次开机后先运行，保持终端打开"],
        ["查看 Orin IP", "hostname -I", "Windows 浏览器需要 http://<orin-ip>:8090/"],
        ["安装免密 sudo", "cd ~/hive_robot/DM_Control_Python\nsudo bash install_hive_robot_sudoers.sh /home/nvidia/hive_robot/DM_Control_Python/left_arm_controller.py", "只需一次；换路径/重装系统后再做"],
        ["测试免密 sudo", "sudo -n python3 /home/nvidia/hive_robot/DM_Control_Python/left_arm_controller.py status", "不提示密码即成功"],
        ["启动网页控制服务", "cd ~/hive_robot/DM_Control_Python\npython3 hp60c_stream_server.py --scan-windows --window-w 80 --window-h 85 --window-step 20 --min-area 30 --max-area 6000 --aim-y-frac 0.25 --target-depth-percentile 75 --port 8090 --enable-controller-execute", "每次测试抓取前运行"],
        ["只看流，不允许网页控制", "python3 hp60c_stream_server.py --scan-windows --window-w 80 --window-h 85 --window-step 20 --min-area 30 --max-area 6000 --aim-y-frac 0.25 --target-depth-percentile 75 --port 8090", "调相机/看画面时用"],
        ["手动执行当前命令", "sudo -n python3 /home/nvidia/hive_robot/DM_Control_Python/left_arm_controller.py target-shoulder --forward <F> --left <L> --up <U> --geometry-execution-blend 1.00 --execute --allow-unverified-geometry", "网页复制命令后可粘贴到终端"],
        ["手动回 Home", "sudo -n python3 /home/nvidia/hive_robot/DM_Control_Python/left_arm_controller.py home", "网页 Home 无响应时备用"],
        ["查看最近训练日志", "tail -n 20 /home/nvidia/hive_robot/DM_Control_Python/data/hp60c_web_runs.jsonl", "把最近记录发给 Codex 分析"],
        ["备份训练日志", "cp /home/nvidia/hive_robot/DM_Control_Python/data/hp60c_web_runs.jsonl ~/hp60c_web_runs_$(date +%Y%m%d_%H%M%S).jsonl", "一天测试结束后备份"],
    ]
    for row in commands:
        ws.append(row)
    style_sheet(ws, [24, 100, 42], 70)

    ws = wb.create_sheet("网页按钮")
    ws.append(["按钮 / 区域", "作用", "建议使用顺序", "风险/备注"])
    buttons = [
        ["skip claw", "勾选后执行目标会加 --skip-claw，只移动到预抓姿态，不闭爪。", "第一次验证姿态时先勾选", "不影响 Home"],
        ["复制命令", "复制当前识别目标生成的 target-shoulder 命令。", "需要手动终端执行或给 Codex 看命令时使用", "命令根据当前画面实时变化"],
        ["执行当前目标", "把当前 shoulder_frame_cm 坐标发送给 left_arm_controller.py。", "先勾 skip claw 验证；姿态正确后取消 skip claw 正式抓取", "需要 --enable-controller-execute 和免密 sudo"],
        ["Home", "执行 left_arm_controller.py home，让机械臂回安全 home。", "每轮测试结束或姿态异常时使用", "同样需要执行权限"],
        ["记录完美", "追加 annotation，标记最近一次执行为 perfect。", "完美抓取后点击", "可在备注里写“完美抓取”"],
        ["记录成功", "追加 annotation，标记最近一次执行为 success。", "抓到但姿态还能优化时点击", "备注写需要优化的关节"],
        ["记录失败", "追加 annotation，标记最近一次执行为 fail。", "失败后点击", "备注越具体越好：手腕太下/侧摆太内/肘太高等"],
        ["State", "实时显示 target、command、坐标、错误信息。", "调试时观察", "数据也会进 jsonl 日志"],
    ]
    for row in buttons:
        ws.append(row)
    style_sheet(ws, [22, 58, 46, 50], 58)

    ws = wb.create_sheet("日志数据")
    ws.append(["项目", "路径 / 内容", "用途"])
    logs = [
        ["默认日志路径", "/home/nvidia/hive_robot/DM_Control_Python/data/hp60c_web_runs.jsonl", "每次网页执行和标注自动追加到这里"],
        ["execute_target 记录", "type=execute_target; target; command; skip_claw; stdout; stderr; returncode; duration_s", "用于分析抓取轨迹、失败原因、重新拟合"],
        ["home 记录", "type=home; command; stdout; stderr; returncode", "记录回 Home 过程"],
        ["annotation 记录", "type=annotation; annotates_run_id; result=perfect/success/fail; note", "把人工观察与执行记录关联起来"],
        ["失败备注模板", "失败：黄圈准，但爪子从物体右侧擦过；侧摆需要向内一点，手腕不变。", "给 Codex 调整局部修正最有用"],
        ["查看最近记录", "tail -n 20 /home/nvidia/hive_robot/DM_Control_Python/data/hp60c_web_runs.jsonl", "直接把输出发给 Codex"],
        ["只导出最近记录", "tail -n 80 /home/nvidia/hive_robot/DM_Control_Python/data/hp60c_web_runs.jsonl > latest_runs.jsonl", "文件太大时只发最近数据"],
    ]
    for row in logs:
        ws.append(row)
    style_sheet(ws, [24, 88, 46], 62)

    ws = wb.create_sheet("故障排查")
    ws.append(["现象", "原因", "处理"])
    trouble = [
        ["网页打不开", "Orin IP 错误、server 未启动、防火墙/网络不通", "Orin 运行 hostname -I；确认 server 打印 listening on 0.0.0.0:8090；Windows 打开 http://<orin-ip>:8090/"],
        ["只有文字没有画面", "首页缓存或 MJPEG 连接异常", "Ctrl+F5 强刷；直接打开 /debug.jpg 和 /stream.mjpg 验证"],
        ["debug.jpg 有图，首页黑屏", "浏览器缓存旧 HTML", "打开 http://<orin-ip>:8090/?v=2 或 Ctrl+F5"],
        ["紫框/黄圈没套目标", "目标颜色/深度检测没有命中；当前逻辑偏黑色目标", "调整目标位置/光照；必要时记录失败备注并发 jsonl 给 Codex"],
        ["点击执行但机械臂不动", "未启用 --enable-controller-execute 或 sudo 免密未配置", "确认启动命令含 --enable-controller-execute；运行 sudo -n python3 ... status 测试"],
        ["终端出现 sudo password", "sudoers 未安装或 controller 路径不匹配", "重新运行 install_hive_robot_sudoers.sh，确保路径是 /home/nvidia/hive_robot/DM_Control_Python/left_arm_controller.py"],
        ["Unable to register with master", "HP60C ROS driver/ROS master 未启动", "先运行 cd ~/ascam_ws; source ~/.bashrc; roslaunch ascamera hp60c.launch"],
        ["cv_bridge 报错", "Orin 上 cv_bridge/OpenCV 环境冲突", "当前 hp60c_stream_server.py 已手动解析 ROS Image，通常可忽略 cv_bridge unavailable 提示"],
        ["libgomp static TLS", "OpenCV 加载顺序问题", "当前脚本已最前 import cv2；如仍失败，用 LD_PRELOAD=/lib/aarch64-linux-gnu/libgomp.so.1 python3 ..."],
        ["日志没有生成", "未执行目标/Home，或当前目录不同", "看 server 启动时打印的 Run log 绝对路径；默认在 DM_Control_Python/data/hp60c_web_runs.jsonl"],
    ]
    for row in trouble:
        ws.append(row)
    style_sheet(ws, [34, 50, 86], 58)

    ws = wb.create_sheet("文件版本")
    ws.append(["字段", "内容"])
    version_rows = [
        ["生成日期", "2026-07-13"],
        ["适用目录", "~/hive_robot/DM_Control_Python"],
        ["主要脚本", "hp60c_stream_server.py, left_arm_controller.py, install_hive_robot_sudoers.sh"],
        ["网页地址", "http://<orin-ip>:8090/"],
        ["控制模式", "target-shoulder, geometry-execution-blend=1.00"],
        ["默认日志", "data/hp60c_web_runs.jsonl"],
    ]
    for row in version_rows:
        ws.append(row)
    style_sheet(ws, [22, 90], 34)

    ws = wb.create_sheet("YOLO on Orin")
    ws.append(["Purpose", "Command / Setting", "When to Use", "Success Signal"])
    yolo_rows = [
        [
            "Check Jetson Docker runtime",
            'sudo docker run --rm --runtime nvidia nvcr.io/nvidia/l4t-base:35.4.1 bash -lc "ls /dev/nvhost* | head"',
            "Run once after Docker/NVIDIA runtime setup",
            "Container prints /dev/nvhost-as-gpu and other /dev/nvhost* devices",
        ],
        [
            "Create YOLO output directory",
            "mkdir -p ~/hive_robot/DM_Control_Python/yolo_runs",
            "Before running YOLO predictions",
            "Directory exists on the Orin host",
        ],
        [
            "Start YOLO JetPack 5 container",
            "sudo docker run -it --rm \\\n  --runtime nvidia \\\n  --network host \\\n  --ipc host \\\n  -v ~/hive_robot:/workspace/hive_robot \\\n  ultralytics/ultralytics:latest-jetson-jetpack5",
            "Use this on Orin JetPack 5 / L4T R35.x",
            "Shell prompt changes to root@ubuntu:/ultralytics#",
        ],
        [
            "Verify CUDA and model loading",
            "python3 - <<'PY'\nimport torch\nfrom ultralytics import YOLO\nprint('torch:', torch.__version__)\nprint('cuda:', torch.cuda.is_available())\nif torch.cuda.is_available():\n    print('gpu:', torch.cuda.get_device_name(0))\nmodel = YOLO('yolo11n.pt')\nprint('YOLO loaded')\nPY",
            "Inside the YOLO container",
            "Shows cuda: True, gpu: Orin, YOLO loaded",
        ],
        [
            "Test built-in COCO image",
            "yolo predict model=yolo11n.pt source=/ultralytics/ultralytics/assets/bus.jpg imgsz=640 conf=0.25",
            "Use this to confirm YOLO itself works",
            "Detects 4 persons and 1 bus",
        ],
        [
            "Capture current HP60C frame",
            "ts=$(date +%s%3N)\nimg=/tmp/hp60c_${ts}.jpg\nwget -O \"$img\" \"http://127.0.0.1:8090/debug.jpg?t=${ts}\"",
            "Use timestamp to avoid cached debug.jpg",
            "Creates a fresh /tmp/hp60c_<timestamp>.jpg",
        ],
        [
            "Detect bottle and scissors on HP60C frame",
            "yolo predict model=yolo11n.pt \\\n  source=\"$img\" \\\n  imgsz=1280 \\\n  conf=0.01 \\\n  classes=39,76 \\\n  project=/workspace/hive_robot/DM_Control_Python/yolo_runs \\\n  name=bottle_scissors_${ts}",
            "Inside the YOLO container after capturing current HP60C frame",
            "Results saved under DM_Control_Python/yolo_runs",
        ],
        [
            "Useful COCO class ids",
            "bottle=39\ncell phone=67\nbook=73\nscissors=76",
            "Use with classes=... to reduce noisy detections",
            "Class-limited detection is cleaner than all-class low-conf detection",
        ],
        [
            "Testing note",
            "Default yolo11n.pt is only a COCO sanity check. HP60C overhead worktable images need class whitelist/per-class thresholds, then a custom HP60C chess/screw/object model.",
            "Use when moving from demo detection to robot control",
            "Scissors detected reliably around 0.41-0.44; bottle was weak around 0.03 in testing",
        ],
    ]
    for row in yolo_rows:
        ws.append(row)
    style_sheet(ws, [28, 92, 44, 50], 74)

    wb.save(OUT)
    load_workbook(OUT)
    print(str(OUT.resolve()).encode("utf-8", errors="replace").decode("ascii", errors="ignore"))


if __name__ == "__main__":
    main()
