from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


out = Path(r"C:\Users\rhche\OneDrive\文档\HiveRobot\左臂电机ID备忘.xlsx")

headers = [
    "Component ID",
    "关节/部位",
    "电机型号",
    "Master ID (Hex)",
    "Master ID (Dec)",
    "Slave ID (Hex)",
    "Slave ID (Dec)",
    "备注",
]

left_arm = [
    [14, "左肩前摆", "DM4340", "0x0E", 14, "0x1E", 30, "图中左臂；已按图编 ID"],
    [15, "左肩侧摆", "DM4340", "0x0F", 15, "0x1F", 31, "图中左臂；已按图编 ID"],
    [16, "左肩旋转", "DM4340", "0x10", 16, "0x20", 32, "图中左臂；已测试 0x20/0x10"],
    [17, "左肘", "DM4340", "0x11", 17, "0x21", 33, "图中左臂；已按图编 ID"],
    [18, "左臂旋转", "DM4340", "0x12", 18, "0x22", 34, "图中左臂；已按图编 ID"],
    [24, "左爪/夹爪", "DM4310", "0x18", 24, "0x28", 40, "后续新增；不在原图规划中；按顺序续编"],
]

full_ref = [
    [1, "左髋旋转", "", "0x01", 1, "0x11", 17, "左腿"],
    [2, "左髋侧摆", "", "0x02", 2, "0x12", 18, "左腿"],
    [3, "左髋前摆", "", "0x03", 3, "0x13", 19, "左腿"],
    [4, "左膝", "", "0x04", 4, "0x14", 20, "左腿"],
    [5, "左脚上", "", "0x05", 5, "0x15", 21, "左脚"],
    [6, "左脚下", "", "0x06", 6, "0x16", 22, "左脚"],
    [7, "右髋旋转", "", "0x07", 7, "0x17", 23, "右腿"],
    [8, "右髋侧摆", "", "0x08", 8, "0x18", 24, "右腿"],
    [9, "右髋前摆", "", "0x09", 9, "0x19", 25, "右腿"],
    [10, "右膝", "", "0x0A", 10, "0x1A", 26, "右腿"],
    [11, "右脚上", "", "0x0B", 11, "0x1B", 27, "右脚"],
    [12, "右脚下", "", "0x0C", 12, "0x1C", 28, "右脚"],
    [13, "躯干旋转", "", "0x0D", 13, "0x1D", 29, "躯干"],
    [14, "左肩前摆", "DM4340", "0x0E", 14, "0x1E", 30, "左臂"],
    [15, "左肩侧摆", "DM4340", "0x0F", 15, "0x1F", 31, "左臂"],
    [16, "左肩旋转", "DM4340", "0x10", 16, "0x20", 32, "左臂"],
    [17, "左肘", "DM4340", "0x11", 17, "0x21", 33, "左臂"],
    [18, "左臂旋转", "DM4340", "0x12", 18, "0x22", 34, "左臂"],
    [19, "右肩前摆", "", "0x13", 19, "0x23", 35, "右臂"],
    [20, "右肩侧摆", "", "0x14", 20, "0x24", 36, "右臂"],
    [21, "右肩旋转", "", "0x15", 21, "0x25", 37, "右臂"],
    [22, "右肘", "", "0x16", 22, "0x26", 38, "右臂"],
    [23, "右臂旋转", "", "0x17", 23, "0x27", 39, "右臂"],
    [24, "左爪/夹爪", "DM4310", "0x18", 24, "0x28", 40, "新增左爪；不在原图规划中；按顺序续编"],
]


def style_sheet(ws, table_name):
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")

    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    widths = [14, 16, 14, 16, 16, 15, 15, 42, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


wb = Workbook()

ws = wb.active
ws.title = "左臂ID备忘"
ws.append(headers)
for row in left_arm:
    ws.append(row)
style_sheet(ws, "LeftArmIdTable")

ws2 = wb.create_sheet("全身参考")
ws2.append(headers)
for row in full_ref:
    ws2.append(row)
style_sheet(ws2, "FullReferenceTable")
for row in ws2.iter_rows(min_row=2):
    if row[7].value == "左臂":
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="E2F0D9")
    if row[1].value == "左爪/夹爪":
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="FFF2CC")

ws3 = wb.create_sheet("测试记录")
ws3.append(["日期", "接口", "电机/关节", "Master ID", "Slave ID", "测试内容", "结果", "备注"])
ws3.append(
    [
        "2026-06-11",
        "USB-CAN / can0",
        "左肩旋转",
        "0x10",
        "0x20",
        "读参/状态/小幅控制",
        "通过",
        "红×但状态反馈和控制正常，低负载可继续观察",
    ]
)
ws3.append(
    [
        "2026-06-11",
        "USB-CAN",
        "左肩前摆",
        "0x0E",
        "0x1E",
        "读参/状态/小幅控制",
        "通过",
        "与 0x20 成对测试已动",
    ]
)
style_sheet(ws3, "TestLogTable")

wb.save(out)
print("created xlsx")
