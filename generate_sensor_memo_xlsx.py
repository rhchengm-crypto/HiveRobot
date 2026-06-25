from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


out = Path(__file__).with_name("HiveRobot_传感器连接测试备忘.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Modules"

headers = [
    "模块",
    "用途",
    "接口/总线",
    "Pi/转接器识别",
    "接线",
    "关键参数",
    "已测成功结果",
    "封装输出建议",
    "注意事项",
]

rows = [
    [
        "ZE07-CO 一氧化碳",
        "CO 中毒风险监测",
        "UART TTL",
        "/dev/ttyAMA0",
        "VCC/Vin->Pi 5V; GND->Pi GND; TXD->Pi GPIO15 pin10/RXD; RXD->Pi GPIO14 pin8/TXD",
        "9600 baud; 9-byte active upload frame",
        "CO 0.5 ppm; range 500.0 ppm; raw ff 04 03 01 00 05 13 88 58",
        '{"sensor":"ze07_co","co_ppm":0.5,"range_ppm":500.0}',
        "/dev/serial0 指向 ttyAMA10，不用于 40Pin UART；使用 /dev/ttyAMA0。",
    ],
    [
        "MAX6675 K 型热电偶",
        "高温/热源监测",
        "SPI",
        "/dev/spidev0.0",
        "GND->Pi GND pin6; VCC->Pi 3.3V pin1; SCK->GPIO11 pin23; CS->GPIO8 pin24; SO->GPIO9 pin21",
        "SPI bus0 CE0; mode 0; 500000 Hz",
        "约 29-30 C；raw 0x03c0/0x03c8 等正常",
        '{"sensor":"max6675","temperature_c":29.5}',
        "VCC 用 3.3V；热电偶端子正负接反会导致读数异常。",
    ],
    [
        "YDLIDAR X3 Pro",
        "2D 激光雷达/空间轮廓",
        "USB Serial CP210x",
        "/dev/ttyUSB0",
        "雷达 USB 转接板/线 -> Pi USB；建议带供电 USB Hub",
        "115200 baud; one-way=yes; sample_rate=5K; scan_frequency≈8Hz",
        "tri_test 成功；Scan received 352-360 points",
        '{"sensor":"ydlidar_x3_pro","points":[...],"scan_hz":8.0}',
        "后续用 by-id 固定设备名；Jetson 上可跑 ROS2/RViz，Pi 可 ser2net 转发。",
    ],
    [
        "LD6002B 60GHz 3D 人体存在雷达",
        "人体存在/3D 位置",
        "USB Serial CP2104",
        "/dev/ttyUSB0",
        "LD6002B 测试底板 Type-C -> Pi USB",
        "115200 baud; 0x0A04 target list; 0x0A0A area presence",
        "presence=True; areas=[1,0,0,0]; targets=1; x/y/z/dop/id 正常",
        '{"sensor":"ld6002b","presence":true,"areas":[1,0,0,0],"targets":[{"x":-0.70,"y":-0.52,"z":1.01,"dop":6,"id":1}]}',
        "0x0A04 len=4 data=00 00 00 00 表示当前帧 targets=0；是否有人优先看 presence/areas。",
    ],
    [
        "USB 红外摄像头",
        "红外图像采集",
        "USB UVC",
        "/dev/video0; /dev/video1",
        "USB 摄像头 -> Pi USB",
        "fswebcam/OpenCV; /dev/video0 为实际视频节点",
        "fswebcam 抓拍成功",
        '{"sensor":"ir_camera","device":"/dev/video0","frame_path":"..."}',
        "可低频抓拍或实时 MJPEG/RTSP/WebSocket 推送给 Jetson。",
    ],
    [
        "PMS5003/PMS5303 粉尘",
        "PM1.0/PM2.5/PM10",
        "UART TTL via CH340 USB-TTL",
        "/dev/ttyUSB0; by-id: usb-1a86_USB_Serial-if00-port0",
        "PMS VCC/5V->USB-TTL 5V; GND->GND; PMS TXD->USB-TTL RXD; PMS RXD->USB-TTL TXD",
        "9600 baud; frame header 42 4d; wake 42 4d e4 00 01 01 75; active 42 4d e1 00 01 01 72",
        "PM1.0=0; PM2.5=0; PM10=0; 0.3um≈148-186; 0.5um≈117-153",
        '{"sensor":"pms5003","pm1_0":0,"pm2_5":0,"pm10":0,"p03":162,"p05":131}',
        "TX/RX 必须交叉；风扇转说明供电正常；需要主动模式命令后开始输出。",
    ],
    [
        "SCD40 CO2 温湿度",
        "CO2/温湿度/密闭通风风险",
        "I2C via CH347T USB-I2C",
        "USB HID 1a86:55dc; hidraw0/hidraw1",
        "CH347T 3.3V->SCD40 VDD; GND->GND; SCL->SCL; SDA->SDA",
        "I2C addr 0x62; ch347api; start 0x21b1; read_measurement 0xec05",
        "CO2=788-789 ppm; temp=24.7-24.8 C; humidity=46.2-47.3%; serial read ok",
        '{"sensor":"scd40","co2_ppm":789,"temp_c":24.7,"humidity":47.3}',
        "SCD40 用 3.3V，不要 5V；CH347T 是 HID 模式，不是 /dev/i2cX。",
    ],
]

ws.append(headers)
for row in rows:
    ws.append(row)

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9E2F3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border

for idx, width in enumerate([22, 20, 22, 30, 52, 46, 42, 60, 48], start=1):
    ws.column_dimensions[get_column_letter(idx)].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
table_ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
table = Table(displayName="SensorModules", ref=table_ref)
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
ws.add_table(table)

bus = wb.create_sheet("BusPlan")
bus_headers = ["类别", "已用设备", "推荐路径/端口", "规划建议"]
bus_rows = [
    ["40Pin UART", "ZE07-CO", "/dev/ttyAMA0", "保留给关键低频传感器；不要用 /dev/serial0。"],
    ["SPI", "MAX6675", "/dev/spidev0.0", "多个 SPI 设备可共用 SCK/MISO/MOSI，单独 CS。"],
    ["USB Serial", "YDLIDAR, LD6002B, PMS5003", "/dev/serial/by-id/...", "必须用 by-id 固定设备，避免 ttyUSB 顺序变化。"],
    ["USB Video", "红外摄像头", "/dev/video0", "可用 OpenCV/ffmpeg/MJPEG 推流。"],
    ["USB-I2C", "SCD40 via CH347T", "hidraw + ch347api", "CH347T 不走 /dev/i2cX；封装时单独适配。"],
    ["未来 RS485", "W01 CO2 / 工业 LEL", "USB-RS485 A/B", "多个 Modbus 设备可共总线，地址不能冲突。"],
]
bus.append(bus_headers)
for row in bus_rows:
    bus.append(row)
for cell in bus[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for row in bus.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
for idx, width in enumerate([18, 32, 34, 62], start=1):
    bus.column_dimensions[get_column_letter(idx)].width = width
bus.freeze_panes = "A2"
bus.auto_filter.ref = bus.dimensions

next_sheet = wb.create_sheet("NextSteps")
next_rows = [
    ["步骤", "内容"],
    ["1", "为所有 USB 串口设备记录 /dev/serial/by-id 路径，写入配置文件。"],
    ["2", "建立 sensors/ 目录：ze07_co.py, max6675_temp.py, pms5003.py, scd40_ch347.py, ld6002b.py, ydlidar_x3.py, ir_camera.py。"],
    ["3", "统一每个模块 read() 返回：ts, sensor, status, data, raw/debug。"],
    ["4", "Pi 端 gateway.py 聚合低频数据为 JSON，通过 MQTT/WebSocket/TCP 发给 Jetson。"],
    ["5", "摄像头单独提供 MJPEG/RTSP 或按需抓拍；YDLIDAR 可先 ser2net 转给 Jetson ROS2。"],
    ["6", "Jetson 端做数据融合、告警阈值、可视化、ROS2/AI 分析。"],
]
for row in next_rows:
    next_sheet.append(row)
for cell in next_sheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for row in next_sheet.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
next_sheet.column_dimensions["A"].width = 12
next_sheet.column_dimensions["B"].width = 90
next_sheet.freeze_panes = "A2"

readme = wb.create_sheet("Readme")
readme.append(["HiveRobot 传感器连接测试备忘"])
readme.append(["用途", "记录截至 2026-06-11 已在 Pi 5 上验证成功的传感器接线、参数、测试结果和封装建议。"])
readme.append(["注意", "安全相关传感器用于机器人辅助监测，不替代认证报警器或生命安全设备。"])
readme["A1"].font = Font(bold=True, size=14)
readme.column_dimensions["A"].width = 16
readme.column_dimensions["B"].width = 100
for row in readme.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(out)
print("xlsx generated")
